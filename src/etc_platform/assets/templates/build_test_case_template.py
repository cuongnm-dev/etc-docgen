"""
build_test_case_template.py — Industry-grade test case xlsx generator.

Builds test-case.xlsx from scratch (programmatic, reproducible, version-controlled).
Replaces the manually-edited BM.QT.04.04 template.

Standards referenced:
  - IEEE 829-2008  / ISO/IEC/IEEE 29119-3 (Test Documentation)
  - ISTQB Foundation Level (test types, techniques)
  - TestRail / Zephyr / Xray column conventions (multi-cycle execution)
  - ETC BM.QT.04.04 (form code preserved on Cover sheet)
  - VN gov compliance: Vietnamese language, NĐ 30/2020 typography hints

Output: 8 sheets — Cover / Instructions / Overview / Dashboard / Tên chức năng /
        Tên API / Danh sách lỗi / Test Data / Lookups (hidden)

Tên chức năng sheet has 25 columns including:
  - Identity: TC ID, Module, Type, Title, Role
  - Risk: Priority, Severity, Linked AC, Tags
  - Execution input: Preconditions, Test Data, Steps, Expected, Audit Evidence
  - Multi-cycle: Cycle 1 + Cycle 2 (each: Status / Actual / Bug ID / Tester / Date)
  - Notes

Run: python build_test_case_template.py [--out PATH]
"""
from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Side, Font, PatternFill, NamedStyle, Color
)
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, ColorScaleRule
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ─────────── Color palette (consistent with VN gov + ETC brand) ───────────
NAVY = "1F4E78"         # Primary headers
LIGHT_BLUE = "D9E1F2"   # Section headers
PALE_BLUE = "EAF1F8"    # Alt row tint
GREEN_PASS = "C6EFCE"; GREEN_TEXT = "006100"
RED_FAIL = "FFC7CE";   RED_TEXT = "9C0006"
YELLOW_BLK = "FFEB9C"; YELLOW_TEXT = "9C5700"
GRAY_SKIP = "EDEDED";  GRAY_TEXT = "595959"
ORANGE_NOT = "FFE4C4"; ORANGE_TEXT = "974706"
RED_CRIT = "DC2626"
ORANGE_HIGH = "EA580C"
YELLOW_MED = "EAB308"
GREEN_LOW = "16A34A"


# ─────────── Style factory (consistent typography) ───────────
def _border(thin=True):
    s = Side(style="thin" if thin else "medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_style(fill=NAVY, font_color="FFFFFF", size=11, bold=True):
    return {
        "font": Font(name="Times New Roman", size=size, bold=bold, color=font_color),
        "fill": PatternFill("solid", fgColor=fill),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _border(),
    }

def _data_style(wrap=True, vertical="top", size=11):
    return {
        "font": Font(name="Times New Roman", size=size),
        "fill": PatternFill(fill_type=None),
        "alignment": Alignment(horizontal="left", vertical=vertical, wrap_text=wrap),
        "border": _border(),
    }

def _apply(cell, style):
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]


# ─────────── Lookup values (single source of truth) ───────────
LOOKUPS = {
    "modules": ["m01","m02","m03","m04","m05","m06","m07","m08","m09","m10","m11","m12","m13"],
    "test_types": ["Smoke","Functional","Regression","Boundary","Negative","Security","Performance","UX","Compatibility","Audit"],
    "priorities": ["Rất cao","Cao","Trung bình","Thấp"],
    "severities": ["Blocker","Major","Minor","Trivial"],
    "statuses_run": ["Pass","Fail","Blocked","Skip","Not-Run"],
    "statuses_track": ["Open","In-Progress","Closed","Re-Open","Resolved"],
    "defect_severities": ["Blocker","Major","Minor","Trivial"],
    "defect_statuses": ["Open","Assigned","In-Progress","Fixed","Verified","Closed","Reopened","Won't Fix"],
}


# ─────────── Sheet builders ───────────
def build_cover(wb: Workbook, project: dict):
    ws = wb.create_sheet("Trang bìa", 0)
    ws.sheet_view.showGridLines = False

    # Column widths
    for col, w in zip("ABCDEFGH", [4, 22, 30, 18, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = w

    # Title block
    ws.merge_cells("B2:H4")
    c = ws["B2"]; c.value = project.get("display_name", "[TÊN HỆ THỐNG]")
    c.font = Font(name="Times New Roman", size=20, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B5:H5")
    c = ws["B5"]; c.value = "KỊCH BẢN KIỂM THỬ — TEST CASE DOCUMENT"
    c.font = Font(name="Times New Roman", size=14, italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B7:H7")
    c = ws["B7"]; c.value = f"Mã biểu mẫu: BM.QT.04.04 · Phiên bản: {project.get('version','1.0')} · Ngày: {project.get('today','dd/mm/yyyy')}"
    c.font = Font(name="Times New Roman", size=10, italic=True, color="666666")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Project info table
    ws["B10"].value = "THÔNG TIN DỰ ÁN"
    ws.merge_cells("B10:H10")
    _apply(ws["B10"], _header_style(fill=LIGHT_BLUE, font_color=NAVY, size=12))
    info_rows = [
        ("Tên dự án", project.get("display_name","")),
        ("Mã dự án", project.get("code","")),
        ("Khách hàng", project.get("client","")),
        ("Đơn vị thực hiện", project.get("dev_unit","")),
        ("Phạm vi mô tả", project.get("description","")),
        ("Công nghệ sử dụng", project.get("tech_stack","")),
    ]
    for i, (k, v) in enumerate(info_rows, start=11):
        ws[f"B{i}"] = k; ws[f"C{i}"] = v
        ws.merge_cells(f"C{i}:H{i}")
        _apply(ws[f"B{i}"], _data_style(wrap=False))
        ws[f"B{i}"].font = Font(name="Times New Roman", size=11, bold=True)
        _apply(ws[f"C{i}"], _data_style(wrap=True))

    # Change history (rows 18+)
    ws["B18"].value = "LỊCH SỬ THAY ĐỔI"
    ws.merge_cells("B18:H18")
    _apply(ws["B18"], _header_style(fill=LIGHT_BLUE, font_color=NAVY, size=12))
    headers = ["Ngày", "Phiên bản", "Loại", "Phạm vi", "Người sửa", "Người duyệt", "Ghi chú"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(19, i, h)
        _apply(c, _header_style())
    # First entry
    initial = [project.get("today",""), project.get("version","1.0"), "Tạo mới", "Toàn bộ tài liệu",
               project.get("dev_unit",""), "[CẦN BỔ SUNG]", ""]
    for i, v in enumerate(initial, start=2):
        c = ws.cell(20, i, v)
        _apply(c, _data_style(wrap=True))
    ws.row_dimensions[19].height = 24
    ws.row_dimensions[20].height = 22


def build_instructions(wb: Workbook):
    ws = wb.create_sheet("Hướng dẫn")
    ws.sheet_view.showGridLines = False
    for col, w in zip("AB", [4, 110]):
        ws.column_dimensions[col].width = w

    title = ws["B2"]; title.value = "HƯỚNG DẪN SỬ DỤNG TÀI LIỆU"
    title.font = Font(name="Times New Roman", size=16, bold=True, color=NAVY)

    sections = [
        ("Mục đích", "Tài liệu chứa toàn bộ kịch bản kiểm thử (test case) của hệ thống. Tester thực hiện theo các bước, ghi lại kết quả thực tế và liên kết bug nếu phát hiện."),
        ("Cấu trúc tài liệu", None),
        ("• Cover", "Thông tin dự án + lịch sử thay đổi tài liệu."),
        ("• Overview", "Phạm vi kiểm thử, môi trường, tiêu chí entry/exit."),
        ("• Dashboard", "Tổng hợp kết quả kiểm thử (auto-update bằng công thức từ các sheet TC)."),
        ("• Tên chức năng", "Test case giao diện — tester thao tác qua trình duyệt theo các bước."),
        ("• Tên API", "Test case API — tester gọi endpoint theo specification."),
        ("• Danh sách lỗi", "Danh sách lỗi phát hiện. Mỗi bug có ID, link tới TC, severity, status."),
        ("• Test Data", "Thư viện dữ liệu test (boundary values, fixtures)."),
        ("• Lookups", "Danh sách giá trị dropdown (ẩn — không sửa)."),
        ("Quy trình thực hiện", None),
        ("Bước 1", "Tester đọc Overview → hiểu phạm vi, môi trường."),
        ("Bước 2", "Mở UI/Tên API → chọn TC theo Module/Priority → đọc Tiền điều kiện + Steps."),
        ("Bước 3", "Thực hiện theo Steps, ghi Kết quả thực tế vào cột tương ứng."),
        ("Bước 4", "Chọn Trạng thái: Pass / Fail / Blocked / Skip / Not-Run từ dropdown."),
        ("Bước 5", "Nếu Fail → log vào Danh sách lỗi → ghi Bug ID vào TC, gắn Severity."),
        ("Bước 6", "Hoàn tất 1 cycle → Dashboard tự cập nhật pass rate, defect breakdown."),
        ("Quy ước Mã TC", "Format: TC-{MOD}-{ROLE}-{DIM}-{NNN}. VD: TC-M02-PHAN_CONG-HAPPY-001."),
        ("Mã ưu tiên (Priority)", "Rất cao = smoke/auth/security/audit. Cao = validation/RBAC. Trung bình = boundary/alternate. Thấp = cosmetic."),
        ("Mức nghiêm trọng (Severity)", "Blocker = hệ thống không dùng được. Major = chức năng chính lỗi. Minor = chức năng phụ. Trivial = cosmetic/typo."),
    ]
    row = 4
    for k, v in sections:
        if v is None:
            c = ws[f"B{row}"]; c.value = k
            c.font = Font(name="Times New Roman", size=13, bold=True, color=NAVY)
            ws.row_dimensions[row].height = 24
        else:
            c = ws[f"B{row}"]
            c.value = f"{k}: {v}" if k.startswith("•") or k.startswith("Bước") or k.startswith("Quy ") or k.startswith("Mã") or k.startswith("Mức") else v
            if not k.startswith(("•", "Bước", "Quy ", "Mã", "Mức")):
                # First-line label cell + body
                c.value = v
                c.font = Font(name="Times New Roman", size=11)
            else:
                c.font = Font(name="Times New Roman", size=11)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1


def build_overview(wb: Workbook, project: dict):
    ws = wb.create_sheet("Tổng quan")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [4, 28, 22, 18, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = w

    ws["B2"].value = "TỔNG QUAN KIỂM THỬ"
    ws.merge_cells("B2:H2")
    ws["B2"].font = Font(name="Times New Roman", size=16, bold=True, color=NAVY)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    sections = [
        ("Phạm vi kiểm thử", project.get("description","[CẦN BỔ SUNG]")),
        ("Đơn vị phát triển", project.get("dev_unit","")),
        ("Khách hàng", project.get("client","")),
        ("Phiên bản hệ thống", project.get("version","1.0")),
        ("Công nghệ", project.get("tech_stack","")),
        ("Môi trường kiểm thử", project.get("test_env","Web + Browser (Chrome/Edge), localhost Docker")),
        ("Thời gian kiểm thử", project.get("test_period","[CẦN BỔ SUNG]")),
        ("Tester chính", project.get("primary_tester","[CẦN BỔ SUNG]")),
        ("Tiêu chí Entry", "Hệ thống đã build + deploy lên môi trường test; có test data đầy đủ; tester đã được training."),
        ("Tiêu chí Exit", "100% TC priority Rất cao + Cao đã thực hiện; pass rate ≥ 95%; không còn defect Blocker/Major mở."),
        ("Test technique áp dụng", "ISTQB: Equivalence Partition, Boundary Value Analysis, Decision Table, State Transition, Error Guessing."),
        ("Tham chiếu", "TKCS, TKCT, HDSD của hệ thống. Acceptance criteria nằm trong feature-catalog.json."),
    ]
    for i, (k, v) in enumerate(sections, start=4):
        ws[f"B{i}"] = k; ws[f"C{i}"] = v
        ws.merge_cells(f"C{i}:H{i}")
        _apply(ws[f"B{i}"], _data_style(wrap=False))
        ws[f"B{i}"].font = Font(name="Times New Roman", size=11, bold=True)
        ws[f"B{i}"].fill = PatternFill("solid", fgColor=PALE_BLUE)
        _apply(ws[f"C{i}"], _data_style(wrap=True))
        ws.row_dimensions[i].height = max(22, 18 * (1 + len(str(v)) // 90))


def build_dashboard(wb: Workbook):
    ws = wb.create_sheet("Tổng hợp kết quả")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJ", [4, 24, 14, 14, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    # Title
    ws["B2"].value = "BẢNG TỔNG HỢP KẾT QUẢ KIỂM THỬ"
    ws.merge_cells("B2:J2")
    _apply(ws["B2"], _header_style(fill=NAVY, font_color="FFFFFF", size=14))
    ws.row_dimensions[2].height = 32

    # ─── Section 1: Overall ───
    ws["B4"].value = "1. KẾT QUẢ TỔNG"; ws.merge_cells("B4:J4")
    _apply(ws["B4"], _header_style(fill=LIGHT_BLUE, font_color=NAVY, size=12))
    headers = ["Phạm vi", "Tổng TC", "Pass", "Fail", "Blocked", "Skip", "Not-Run", "Pass rate", "Tỷ lệ thực hiện"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(5, i, h); _apply(c, _header_style())
    ws.row_dimensions[5].height = 28

    # UI row
    ws.cell(6, 2, "Tên chức năng")
    ws.cell(6, 3, '=COUNTA(\'Tên chức năng\'!A12:A1000)')
    ws.cell(6, 4, '=COUNTIF(\'Tên chức năng\'!O12:O1000,"Pass")+COUNTIF(\'Tên chức năng\'!T12:T1000,"Pass")')
    ws.cell(6, 5, '=COUNTIF(\'Tên chức năng\'!O12:O1000,"Fail")+COUNTIF(\'Tên chức năng\'!T12:T1000,"Fail")')
    ws.cell(6, 6, '=COUNTIF(\'Tên chức năng\'!O12:O1000,"Blocked")+COUNTIF(\'Tên chức năng\'!T12:T1000,"Blocked")')
    ws.cell(6, 7, '=COUNTIF(\'Tên chức năng\'!O12:O1000,"Skip")+COUNTIF(\'Tên chức năng\'!T12:T1000,"Skip")')
    ws.cell(6, 8, '=C6-D6-E6-F6-G6')
    ws.cell(6, 9, '=IFERROR(D6/(D6+E6),0)')
    ws.cell(6, 10, '=IFERROR((D6+E6+F6+G6)/C6,0)')

    # API row
    ws.cell(7, 2, "Tên API")
    ws.cell(7, 3, '=COUNTA(\'Tên API\'!A12:A1000)')
    ws.cell(7, 4, '=COUNTIF(\'Tên API\'!O12:O1000,"Pass")+COUNTIF(\'Tên API\'!T12:T1000,"Pass")')
    ws.cell(7, 5, '=COUNTIF(\'Tên API\'!O12:O1000,"Fail")+COUNTIF(\'Tên API\'!T12:T1000,"Fail")')
    ws.cell(7, 6, '=COUNTIF(\'Tên API\'!O12:O1000,"Blocked")+COUNTIF(\'Tên API\'!T12:T1000,"Blocked")')
    ws.cell(7, 7, '=COUNTIF(\'Tên API\'!O12:O1000,"Skip")+COUNTIF(\'Tên API\'!T12:T1000,"Skip")')
    ws.cell(7, 8, '=C7-D7-E7-F7-G7')
    ws.cell(7, 9, '=IFERROR(D7/(D7+E7),0)')
    ws.cell(7, 10, '=IFERROR((D7+E7+F7+G7)/C7,0)')

    # Total row
    ws.cell(8, 2, "TỔNG"); ws.cell(8, 2).font = Font(name="Times New Roman", size=11, bold=True)
    for col_idx in range(3, 9):
        ws.cell(8, col_idx, f'=SUM({get_column_letter(col_idx)}6:{get_column_letter(col_idx)}7)')
    ws.cell(8, 9, '=IFERROR(D8/(D8+E8),0)')
    ws.cell(8, 10, '=IFERROR((D8+E8+F8+G8)/C8,0)')

    # Style data rows
    for r in [6, 7, 8]:
        for c_idx in range(2, 11):
            cell = ws.cell(r, c_idx)
            _apply(cell, _data_style(wrap=False, vertical="center"))
            if r == 8:
                cell.font = Font(name="Times New Roman", size=11, bold=True)
                cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        ws.row_dimensions[r].height = 22
    # % format on cols I, J
    for r in [6, 7, 8]:
        ws.cell(r, 9).number_format = "0.0%"
        ws.cell(r, 10).number_format = "0.0%"

    # ─── Section 2: By Priority ───
    ws["B11"].value = "2. KẾT QUẢ THEO MỨC ƯU TIÊN"; ws.merge_cells("B11:J11")
    _apply(ws["B11"], _header_style(fill=LIGHT_BLUE, font_color=NAVY, size=12))
    pri_hdr = ["Mức ưu tiên", "Tổng TC", "Pass", "Fail", "Blocked", "Pass rate"]
    for i, h in enumerate(pri_hdr, start=2):
        _apply(ws.cell(12, i, h), _header_style())
    for j, pri in enumerate(LOOKUPS["priorities"]):
        r = 13 + j
        ws.cell(r, 2, pri)
        ws.cell(r, 3, f'=COUNTIF(\'Tên chức năng\'!F12:F1000,"{pri}")+COUNTIF(\'Tên API\'!F12:F1000,"{pri}")')
        ws.cell(r, 4, f'=COUNTIFS(\'Tên chức năng\'!F12:F1000,"{pri}",\'Tên chức năng\'!O12:O1000,"Pass")+COUNTIFS(\'Tên API\'!F12:F1000,"{pri}",\'Tên API\'!O12:O1000,"Pass")')
        ws.cell(r, 5, f'=COUNTIFS(\'Tên chức năng\'!F12:F1000,"{pri}",\'Tên chức năng\'!O12:O1000,"Fail")+COUNTIFS(\'Tên API\'!F12:F1000,"{pri}",\'Tên API\'!O12:O1000,"Fail")')
        ws.cell(r, 6, f'=COUNTIFS(\'Tên chức năng\'!F12:F1000,"{pri}",\'Tên chức năng\'!O12:O1000,"Blocked")+COUNTIFS(\'Tên API\'!F12:F1000,"{pri}",\'Tên API\'!O12:O1000,"Blocked")')
        ws.cell(r, 7, f'=IFERROR(D{r}/(D{r}+E{r}),0)')
        for c_idx in range(2, 8):
            _apply(ws.cell(r, c_idx), _data_style(wrap=False, vertical="center"))
        ws.cell(r, 7).number_format = "0.0%"

    # ─── Section 3: Phân loại lỗi (Defect by Severity) ───
    # 2D crosstab: Status (rows) × Severity (cols) — user-expected layout
    ws["B19"].value = "3. PHÂN LOẠI LỖI"; ws.merge_cells("B19:J19")
    _apply(ws["B19"], _header_style(fill=LIGHT_BLUE, font_color=NAVY, size=12))

    # Sub-header row 20: "Trạng thái lỗi" | "Phân loại lỗi theo severity (Critical/Major/Minor/Trivial)" | "Tổng"
    ws.cell(20, 2, "Trạng thái lỗi"); _apply(ws.cell(20, 2), _header_style())
    ws.merge_cells("C20:F20"); ws.cell(20, 3, "Phân loại lỗi theo severity")
    _apply(ws.cell(20, 3), _header_style())
    ws.cell(20, 7, "Tổng"); _apply(ws.cell(20, 7), _header_style())
    ws.row_dimensions[20].height = 24

    # Severity column headers row 21
    severities = LOOKUPS["defect_severities"]  # Blocker, Major, Minor, Trivial
    ws.cell(21, 2, ""); _apply(ws.cell(21, 2), _header_style(fill=PALE_BLUE, font_color=NAVY))
    for j, sev in enumerate(severities):
        c = ws.cell(21, 3 + j, sev); _apply(c, _header_style(fill=PALE_BLUE, font_color=NAVY))
    ws.cell(21, 7, ""); _apply(ws.cell(21, 7), _header_style(fill=PALE_BLUE, font_color=NAVY))
    ws.row_dimensions[21].height = 22

    # Status rows 22-25: Open / In-Progress / Fixed / Closed
    statuses = ["Open", "In-Progress", "Fixed", "Closed"]
    for i, status in enumerate(statuses):
        r = 22 + i
        ws.cell(r, 2, status); _apply(ws.cell(r, 2), _data_style(wrap=False, vertical="center"))
        ws.cell(r, 2).font = Font(name="Times New Roman", size=11, bold=True)
        for j, sev in enumerate(severities):
            ws.cell(r, 3 + j, f'=COUNTIFS(\'Danh sách lỗi\'!E5:E500,"{sev}",\'Danh sách lỗi\'!I5:I500,"{status}")')
            _apply(ws.cell(r, 3 + j), _data_style(wrap=False, vertical="center"))
        ws.cell(r, 7, f'=SUM(C{r}:F{r})')
        _apply(ws.cell(r, 7), _data_style(wrap=False, vertical="center"))
        ws.cell(r, 7).font = Font(name="Times New Roman", size=11, bold=True)

    # Total row 26: "Mọi trạng thái (không tính Cancelled)"
    r = 26
    ws.cell(r, 2, "Mọi trạng thái")
    _apply(ws.cell(r, 2), _data_style(wrap=False, vertical="center"))
    ws.cell(r, 2).font = Font(name="Times New Roman", size=11, bold=True, italic=True)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=PALE_BLUE)
    for j, sev in enumerate(severities):
        # Sum vertically across status rows (22-25)
        col = get_column_letter(3 + j)
        ws.cell(r, 3 + j, f'=SUM({col}22:{col}25)')
        _apply(ws.cell(r, 3 + j), _data_style(wrap=False, vertical="center"))
        ws.cell(r, 3 + j).font = Font(name="Times New Roman", size=11, bold=True)
        ws.cell(r, 3 + j).fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws.cell(r, 7, '=SUM(C26:F26)')
    _apply(ws.cell(r, 7), _data_style(wrap=False, vertical="center"))
    ws.cell(r, 7).font = Font(name="Times New Roman", size=11, bold=True)
    ws.cell(r, 7).fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws.row_dimensions[r].height = 24

    # ─── Section 4: Tổng hợp theo Module ───
    ws["B28"].value = "4. KẾT QUẢ THEO PHÂN HỆ (MODULE)"; ws.merge_cells("B28:J28")
    _apply(ws["B28"], _header_style(fill=LIGHT_BLUE, font_color=NAVY, size=12))
    mod_hdr = ["Phân hệ", "Tổng TC", "Pass", "Fail", "Blocked", "Pass rate"]
    for i, h in enumerate(mod_hdr, start=2):
        _apply(ws.cell(29, i, h), _header_style())
    for j, mod in enumerate(LOOKUPS["modules"]):
        r = 30 + j
        ws.cell(r, 2, mod)
        ws.cell(r, 3, f'=COUNTIF(\'Tên chức năng\'!B13:B1000,"{mod}")+COUNTIF(\'Tên API\'!B13:B1000,"{mod}")')
        ws.cell(r, 4, f'=COUNTIFS(\'Tên chức năng\'!B13:B1000,"{mod}",\'Tên chức năng\'!O13:O1000,"Pass")+COUNTIFS(\'Tên API\'!B13:B1000,"{mod}",\'Tên API\'!O13:O1000,"Pass")')
        ws.cell(r, 5, f'=COUNTIFS(\'Tên chức năng\'!B13:B1000,"{mod}",\'Tên chức năng\'!O13:O1000,"Fail")+COUNTIFS(\'Tên API\'!B13:B1000,"{mod}",\'Tên API\'!O13:O1000,"Fail")')
        ws.cell(r, 6, f'=COUNTIFS(\'Tên chức năng\'!B13:B1000,"{mod}",\'Tên chức năng\'!O13:O1000,"Blocked")+COUNTIFS(\'Tên API\'!B13:B1000,"{mod}",\'Tên API\'!O13:O1000,"Blocked")')
        ws.cell(r, 7, f'=IFERROR(D{r}/(D{r}+E{r}),0)')
        for c_idx in range(2, 8):
            _apply(ws.cell(r, c_idx), _data_style(wrap=False, vertical="center"))
        ws.cell(r, 7).number_format = "0.0%"


def _build_tc_sheet(wb: Workbook, sheet_name: str, start_data_row: int = 12):
    """Build UI/API test case sheet with 25 industry-grade columns."""
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "B12"  # Freeze header rows + col A

    # Column structure (25 cols)
    cols = [
        ("A", "Mã TC", 16),
        ("B", "Phân hệ", 12),
        ("C", "Loại test", 14),
        ("D", "Mục đích", 38),
        ("E", "Vai trò", 18),
        ("F", "Mức ưu tiên", 13),
        ("G", "Mức nghiêm trọng", 14),
        ("H", "Liên kết AC", 14),
        ("I", "Tags", 18),
        ("J", "Tiền điều kiện", 30),
        ("K", "Dữ liệu test", 22),
        ("L", "Các bước thực hiện", 45),
        ("M", "Kết quả mong đợi", 45),
        ("N", "Bằng chứng (Audit)", 26),
        # Cycle 1
        ("O", "Lần 1 — Trạng thái", 13),
        ("P", "Lần 1 — Kết quả thực tế", 28),
        ("Q", "Lần 1 — Bug ID", 12),
        ("R", "Lần 1 — Người test", 14),
        ("S", "Lần 1 — Ngày", 12),
        # Cycle 2
        ("T", "Lần 2 — Trạng thái", 13),
        ("U", "Lần 2 — Kết quả thực tế", 28),
        ("V", "Lần 2 — Bug ID", 12),
        ("W", "Lần 2 — Người test", 14),
        ("X", "Lần 2 — Ngày", 12),
        ("Y", "Ghi chú", 24),
    ]
    for col, _name, w in cols:
        ws.column_dimensions[col].width = w

    # Title row 1-3
    ws.merge_cells("A1:Y1")
    ws["A1"].value = f"{sheet_name.upper()}"
    _apply(ws["A1"], _header_style(fill=NAVY, font_color="FFFFFF", size=14))
    ws.row_dimensions[1].height = 32

    # Project meta rows 2-4
    meta_labels = [("A2", "Tên hệ thống:"), ("A3", "Phiên bản:"), ("A4", "Người chuẩn bị:")]
    for cell_ref, label in meta_labels:
        ws[cell_ref].value = label
        ws[cell_ref].font = Font(name="Times New Roman", size=11, bold=True)
        ws[cell_ref].alignment = Alignment(horizontal="right")
        # Merge value cell across B:Y for readability — actual value filled by engine
        merge_target = f"B{cell_ref[1]}:Y{cell_ref[1]}"
        ws.merge_cells(merge_target)

    # Summary band row 6-8 (auto-aggregating formulas)
    ws.merge_cells("A6:C6"); ws["A6"].value = "TỔNG HỢP CYCLE 1"
    _apply(ws["A6"], _header_style(fill=LIGHT_BLUE, font_color=NAVY))
    summary_labels_c1 = ["Tổng TC", "Pass", "Fail", "Blocked", "Skip", "Not-Run", "Pass rate"]
    for i, lbl in enumerate(summary_labels_c1):
        ws.cell(7, 4 + i, lbl); _apply(ws.cell(7, 4 + i), _header_style())
    end_data = 1000
    ws.cell(8, 4, f'=COUNTA(A12:A{end_data})')
    ws.cell(8, 5, f'=COUNTIF(O12:O{end_data},"Pass")')
    ws.cell(8, 6, f'=COUNTIF(O12:O{end_data},"Fail")')
    ws.cell(8, 7, f'=COUNTIF(O12:O{end_data},"Blocked")')
    ws.cell(8, 8, f'=COUNTIF(O12:O{end_data},"Skip")')
    ws.cell(8, 9, f'=D8-E8-F8-G8-H8')
    ws.cell(8, 10, f'=IFERROR(E8/(E8+F8),0)')
    ws.cell(8, 10).number_format = "0.0%"
    for c_idx in range(4, 11):
        _apply(ws.cell(8, c_idx), _data_style(wrap=False, vertical="center"))

    ws.merge_cells("A9:C9"); ws["A9"].value = "TỔNG HỢP CYCLE 2"
    _apply(ws["A9"], _header_style(fill=LIGHT_BLUE, font_color=NAVY))
    for i, lbl in enumerate(summary_labels_c1):
        ws.cell(10, 4 + i, lbl); _apply(ws.cell(10, 4 + i), _header_style())
    ws.cell(11, 4, f'=COUNTA(A12:A{end_data})')
    ws.cell(11, 5, f'=COUNTIF(T12:T{end_data},"Pass")')
    ws.cell(11, 6, f'=COUNTIF(T12:T{end_data},"Fail")')
    ws.cell(11, 7, f'=COUNTIF(T12:T{end_data},"Blocked")')
    ws.cell(11, 8, f'=COUNTIF(T12:T{end_data},"Skip")')
    ws.cell(11, 9, f'=D11-E11-F11-G11-H11')
    ws.cell(11, 10, f'=IFERROR(E11/(E11+F11),0)')
    ws.cell(11, 10).number_format = "0.0%"
    for c_idx in range(4, 11):
        _apply(ws.cell(11, c_idx), _data_style(wrap=False, vertical="center"))

    # Header row at row 11 (or use start_data_row - 1)
    # Actually let me put header at row start_data_row - 1, data starts at start_data_row
    header_row = start_data_row - 1  # = 11
    # Wait — row 11 already used by Cycle 2 summary. Let me use start_data_row - 1 = 11 for header.
    # Conflict resolution: put TC header row at row 11, but row 11 has Cycle 2 summary E-J.
    # Solution: extend Cycle 2 summary into row 11 cols A-J only (already done above row 11).
    # Then header row at row 11 won't conflict because we use cols A-Y but cycle summary uses D-J.
    # Actually we DID write Cycle 2 summary at row 11 already. That's wrong placement.
    # Fix: move Cycle 2 summary up, put header at correct row. Let me restructure:
    #  Row 1: title
    #  Row 2-4: project meta
    #  Row 5: blank
    #  Row 6: "TỔNG HỢP CYCLE 1" merge A:C
    #  Row 7: cycle 1 sub-headers D:J
    #  Row 8: cycle 1 values D:J
    #  Row 9: "TỔNG HỢP CYCLE 2" merge A:C
    #  Row 10: cycle 2 sub-headers D:J
    #  Row 11: cycle 2 values D:J
    #  Row (header_row=12): main TC header A:Y    <-- conflict with start_data_row=12
    #
    # Re-decide: start_data_row = 13. Header at row 12.

    # Actually I realize the original code writes to row 11 for Cycle 2 values which conflicts with
    # header row=11. Let me fix start_data_row to 13.
    pass  # Will fix below in new logic


def build_tc_sheet_v2(wb: Workbook, sheet_name: str, start_data_row: int = 13, end_data: int = 1000):
    """Cleaner rebuild: header at row 12, data starts at row 13."""
    ws = wb.create_sheet(sheet_name)

    cols = [
        ("A", "Mã TC", 18),
        ("B", "Phân hệ", 11),
        ("C", "Loại test", 14),
        ("D", "Mục đích", 40),
        ("E", "Vai trò", 20),
        ("F", "Mức ưu tiên", 13),
        ("G", "Nghiêm trọng", 13),
        ("H", "Link AC", 12),
        ("I", "Tags", 18),
        ("J", "Tiền điều kiện", 32),
        ("K", "Dữ liệu test", 22),
        ("L", "Các bước thực hiện", 48),
        ("M", "Kết quả mong đợi", 48),
        ("N", "Bằng chứng audit", 26),
        ("O", "Lần 1 — Trạng thái", 14),
        ("P", "Lần 1 — Kết quả thực tế", 30),
        ("Q", "Lần 1 — Bug ID", 12),
        ("R", "Lần 1 — Người test", 14),
        ("S", "Lần 1 — Ngày", 12),
        ("T", "Lần 2 — Trạng thái", 14),
        ("U", "Lần 2 — Kết quả thực tế", 30),
        ("V", "Lần 2 — Bug ID", 12),
        ("W", "Lần 2 — Người test", 14),
        ("X", "Lần 2 — Ngày", 12),
        ("Y", "Ghi chú", 24),
    ]
    for col, _n, w in cols:
        ws.column_dimensions[col].width = w

    # Row 1 — Title
    ws.merge_cells("A1:Y1")
    ws["A1"].value = sheet_name.upper()
    _apply(ws["A1"], _header_style(fill=NAVY, font_color="FFFFFF", size=14))
    ws.row_dimensions[1].height = 30

    # Row 2-4 — Project metadata (filled by engine)
    meta_labels = ["Tên hệ thống:", "Phiên bản:", "Người chuẩn bị:"]
    for i, lbl in enumerate(meta_labels):
        r = 2 + i
        ws.cell(r, 1, lbl)
        ws.cell(r, 1).font = Font(name="Times New Roman", size=11, bold=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="right")
        ws.merge_cells(f"B{r}:Y{r}")
        ws.cell(r, 2, "")  # value cell to be filled
        _apply(ws.cell(r, 2), _data_style(wrap=False, vertical="center"))

    # Row 6-8 — Cycle 1 summary band
    ws.merge_cells("A6:C6")
    ws["A6"].value = "TỔNG HỢP LẦN 1"
    _apply(ws["A6"], _header_style(fill=LIGHT_BLUE, font_color=NAVY))
    sub_hdrs = ["Tổng", "Pass", "Fail", "Blocked", "Skip", "Not-Run", "Pass rate"]
    for i, h in enumerate(sub_hdrs):
        c = ws.cell(7, 4 + i, h); _apply(c, _header_style())
    ws.cell(8, 4, f'=COUNTA(A{start_data_row}:A{end_data})')
    ws.cell(8, 5, f'=COUNTIF(O{start_data_row}:O{end_data},"Pass")')
    ws.cell(8, 6, f'=COUNTIF(O{start_data_row}:O{end_data},"Fail")')
    ws.cell(8, 7, f'=COUNTIF(O{start_data_row}:O{end_data},"Blocked")')
    ws.cell(8, 8, f'=COUNTIF(O{start_data_row}:O{end_data},"Skip")')
    ws.cell(8, 9, '=D8-E8-F8-G8-H8')
    ws.cell(8, 10, '=IFERROR(E8/(E8+F8),0)')
    ws.cell(8, 10).number_format = "0.0%"
    for ci in range(4, 11):
        _apply(ws.cell(8, ci), _data_style(wrap=False, vertical="center"))
    ws.row_dimensions[7].height = 22; ws.row_dimensions[8].height = 22

    # Row 9-11 — Cycle 2 summary band
    ws.merge_cells("A9:C9")
    ws["A9"].value = "TỔNG HỢP LẦN 2"
    _apply(ws["A9"], _header_style(fill=LIGHT_BLUE, font_color=NAVY))
    for i, h in enumerate(sub_hdrs):
        c = ws.cell(10, 4 + i, h); _apply(c, _header_style())
    ws.cell(11, 4, f'=COUNTA(A{start_data_row}:A{end_data})')
    ws.cell(11, 5, f'=COUNTIF(T{start_data_row}:T{end_data},"Pass")')
    ws.cell(11, 6, f'=COUNTIF(T{start_data_row}:T{end_data},"Fail")')
    ws.cell(11, 7, f'=COUNTIF(T{start_data_row}:T{end_data},"Blocked")')
    ws.cell(11, 8, f'=COUNTIF(T{start_data_row}:T{end_data},"Skip")')
    ws.cell(11, 9, '=D11-E11-F11-G11-H11')
    ws.cell(11, 10, '=IFERROR(E11/(E11+F11),0)')
    ws.cell(11, 10).number_format = "0.0%"
    for ci in range(4, 11):
        _apply(ws.cell(11, ci), _data_style(wrap=False, vertical="center"))
    ws.row_dimensions[10].height = 22; ws.row_dimensions[11].height = 22

    # Row 12 — Main column header
    header_row = start_data_row - 1  # = 12
    for i, (col, name, _w) in enumerate(cols, start=1):
        c = ws.cell(header_row, i, name)
        _apply(c, _header_style())
    ws.row_dimensions[header_row].height = 36

    # Pre-format data region rows (start_data_row .. end_data) with neutral style + uniform height
    data_style_short = _data_style(wrap=False, vertical="center")
    data_style_long = _data_style(wrap=True, vertical="top")
    short_cols = {"A","B","C","E","F","G","H","O","Q","R","S","T","V","W","X"}
    long_cols = {"D","I","J","K","L","M","N","P","U","Y"}
    for r in range(start_data_row, min(end_data + 1, start_data_row + 50)):
        # Pre-format only first 50 rows by default; engine will handle the rest at fill time
        for i, (col, _n, _w) in enumerate(cols, start=1):
            cell = ws.cell(r, i)
            if col in short_cols:
                _apply(cell, data_style_short)
            else:
                _apply(cell, data_style_long)
        ws.row_dimensions[r].height = 36

    # Date format on cycle date columns S, X
    for col in ["S", "X"]:
        for r in range(start_data_row, end_data + 1):
            ws[f"{col}{r}"].number_format = "dd/mm/yyyy"

    # Freeze panes — header + col A frozen
    ws.freeze_panes = ws.cell(start_data_row, 2)  # B<start_data_row>

    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:Y{header_row}"

    # ─── Data validations (dropdowns) ───
    dvs = [
        # (formula1, columns)
        ("='_Tra cứu'!$B$2:$B$20", ["B"]),  # modules
        ("='_Tra cứu'!$C$2:$C$20", ["C"]),  # test_types
        ("='_Tra cứu'!$D$2:$D$20", ["E"]),  # roles
        ("='_Tra cứu'!$E$2:$E$20", ["F"]),  # priorities
        ("='_Tra cứu'!$F$2:$F$20", ["G"]),  # severities
        ("='_Tra cứu'!$G$2:$G$20", ["O", "T"]),  # statuses_run
    ]
    for formula1, dv_cols in dvs:
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        for dc in dv_cols:
            dv.add(f"{dc}{start_data_row}:{dc}{end_data}")
        ws.add_data_validation(dv)

    # ─── Conditional formatting ───
    # Status cells (O, T): green=Pass, red=Fail, yellow=Blocked, gray=Skip, orange=Not-Run
    status_rules = [
        ("Pass", GREEN_PASS, GREEN_TEXT),
        ("Fail", RED_FAIL, RED_TEXT),
        ("Blocked", YELLOW_BLK, YELLOW_TEXT),
        ("Skip", GRAY_SKIP, GRAY_TEXT),
        ("Not-Run", ORANGE_NOT, ORANGE_TEXT),
    ]
    for col in ["O", "T"]:
        rng = f"{col}{start_data_row}:{col}{end_data}"
        for val, bg, fg in status_rules:
            r = CellIsRule(operator="equal", formula=[f'"{val}"'],
                           stopIfTrue=False,
                           font=Font(name="Times New Roman", size=11, bold=True, color=fg),
                           fill=PatternFill("solid", fgColor=bg))
            ws.conditional_formatting.add(rng, r)

    # Priority col F: red if Rất cao, orange Cao, yellow Trung bình, green Thấp
    priority_rules = [
        ("Rất cao", RED_CRIT, "FFFFFF"),
        ("Cao", ORANGE_HIGH, "FFFFFF"),
        ("Trung bình", YELLOW_MED, "1A1A1A"),
        ("Thấp", GREEN_LOW, "FFFFFF"),
    ]
    rng = f"F{start_data_row}:F{end_data}"
    for val, bg, fg in priority_rules:
        r = CellIsRule(operator="equal", formula=[f'"{val}"'], stopIfTrue=False,
                       font=Font(name="Times New Roman", size=11, bold=True, color=fg),
                       fill=PatternFill("solid", fgColor=bg))
        ws.conditional_formatting.add(rng, r)

    # Severity col G similar
    sev_rules = [
        ("Blocker", RED_CRIT, "FFFFFF"),
        ("Major", ORANGE_HIGH, "FFFFFF"),
        ("Minor", YELLOW_MED, "1A1A1A"),
        ("Trivial", GRAY_SKIP, "595959"),
    ]
    rng = f"G{start_data_row}:G{end_data}"
    for val, bg, fg in sev_rules:
        r = CellIsRule(operator="equal", formula=[f'"{val}"'], stopIfTrue=False,
                       font=Font(name="Times New Roman", size=11, bold=True, color=fg),
                       fill=PatternFill("solid", fgColor=bg))
        ws.conditional_formatting.add(rng, r)

    return ws


def build_defect_log(wb: Workbook):
    ws = wb.create_sheet("Danh sách lỗi")
    cols = [
        ("A", "Bug ID", 12),
        ("B", "Tiêu đề", 40),
        ("C", "TC liên quan", 22),
        ("D", "Module", 12),
        ("E", "Severity", 12),
        ("F", "Priority", 12),
        ("G", "Người báo cáo", 16),
        ("H", "Ngày báo", 12),
        ("I", "Trạng thái", 14),
        ("J", "Người sửa", 16),
        ("K", "Ngày sửa", 12),
        ("L", "Mô tả + Cách tái hiện", 50),
        ("M", "Ghi chú", 24),
    ]
    for col, _n, w in cols:
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"].value = "DANH SÁCH LỖI / DEFECT LOG"
    _apply(ws["A1"], _header_style(fill=NAVY, font_color="FFFFFF", size=14))
    ws.row_dimensions[1].height = 30

    # Summary band rows 2-3
    ws.merge_cells("A2:B2"); ws["A2"].value = "Tổng số lỗi:"
    _apply(ws["A2"], _data_style(wrap=False, vertical="center"))
    ws["A2"].font = Font(name="Times New Roman", size=11, bold=True)
    ws["C2"] = '=COUNTA(A5:A1000)'
    _apply(ws["C2"], _data_style(wrap=False, vertical="center"))

    # Header row 4
    for i, (col, name, _w) in enumerate(cols, start=1):
        c = ws.cell(4, i, name); _apply(c, _header_style())
    ws.row_dimensions[4].height = 36

    # Pre-format data rows
    end_data = 500
    data_style_long = _data_style(wrap=True, vertical="top")
    data_style_short = _data_style(wrap=False, vertical="center")
    short_cols = {"A","D","E","F","G","H","I","J","K"}
    for r in range(5, min(end_data + 1, 30)):
        for i, (col, _n, _w) in enumerate(cols, start=1):
            cell = ws.cell(r, i)
            if col in short_cols:
                _apply(cell, data_style_short)
            else:
                _apply(cell, data_style_long)
        ws.row_dimensions[r].height = 32

    # Date columns
    for col in ["H", "K"]:
        for r in range(5, end_data + 1):
            ws[f"{col}{r}"].number_format = "dd/mm/yyyy"

    # Data validations
    dvs = [
        ("='_Tra cứu'!$B$2:$B$20", ["D"]),
        ("='_Tra cứu'!$F$2:$F$20", ["E"]),
        ("='_Tra cứu'!$E$2:$E$20", ["F"]),
        ("='_Tra cứu'!$H$2:$H$20", ["I"]),
    ]
    for formula1, dv_cols in dvs:
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        for dc in dv_cols:
            dv.add(f"{dc}5:{dc}{end_data}")
        ws.add_data_validation(dv)

    # Conditional formatting on Severity (E) + Status (I)
    sev_rules = [
        ("Blocker", RED_CRIT, "FFFFFF"),
        ("Major", ORANGE_HIGH, "FFFFFF"),
        ("Minor", YELLOW_MED, "1A1A1A"),
        ("Trivial", GRAY_SKIP, "595959"),
    ]
    for val, bg, fg in sev_rules:
        r = CellIsRule(operator="equal", formula=[f'"{val}"'], stopIfTrue=False,
                       font=Font(name="Times New Roman", size=11, bold=True, color=fg),
                       fill=PatternFill("solid", fgColor=bg))
        ws.conditional_formatting.add(f"E5:E{end_data}", r)

    status_rules = [
        ("Open", RED_FAIL, RED_TEXT),
        ("In-Progress", YELLOW_BLK, YELLOW_TEXT),
        ("Fixed", LIGHT_BLUE, NAVY),
        ("Verified", GREEN_PASS, GREEN_TEXT),
        ("Closed", GREEN_PASS, GREEN_TEXT),
        ("Reopened", RED_FAIL, RED_TEXT),
        ("Won't Fix", GRAY_SKIP, GRAY_TEXT),
    ]
    for val, bg, fg in status_rules:
        r = CellIsRule(operator="equal", formula=[f'"{val}"'], stopIfTrue=False,
                       font=Font(name="Times New Roman", size=11, bold=True, color=fg),
                       fill=PatternFill("solid", fgColor=bg))
        ws.conditional_formatting.add(f"I5:I{end_data}", r)

    # Freeze panes
    ws.freeze_panes = ws["B5"]
    ws.auto_filter.ref = f"A4:M4"


def build_test_data(wb: Workbook):
    ws = wb.create_sheet("Dữ liệu test")
    for col, w in zip("ABCDE", [4, 30, 40, 22, 22]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:E2")
    ws["B2"].value = "THƯ VIỆN DỮ LIỆU TEST"
    _apply(ws["B2"], _header_style(fill=NAVY, font_color="FFFFFF", size=14))
    ws.row_dimensions[2].height = 30

    sections = [
        ("Tài khoản test", [
            ("Tên đăng nhập", "Mật khẩu", "Vai trò", "Phạm vi"),
            ("admin@test.local", "Test@123", "ADMIN", "Toàn hệ thống"),
            ("hqdk01@test.local", "Test@123", "HQDK", "Cán bộ HQ đăng ký"),
            ("kiemhoa01@test.local", "Test@123", "KIEM_HOA", "Cán bộ kiểm hóa"),
            ("phancong01@test.local", "Test@123", "PHAN_CONG", "Cán bộ phân công"),
            ("lanhdao01@test.local", "Test@123", "LANH_DAO", "Lãnh đạo HQ"),
            ("giamsat01@test.local", "Test@123", "GIAM_SAT", "Giám sát KVGS"),
        ]),
        ("Boundary value templates", [
            ("Trường", "Giá trị min hợp lệ", "Giá trị max hợp lệ", "Giá trị vi phạm"),
            ("Tên hồ sơ", "1 ký tự", "200 ký tự", "0 ký tự / 201 ký tự"),
            ("Mã hồ sơ", "Không có space, A-Z 0-9", "20 ký tự", "Có space / >20 ký tự"),
            ("Số tiền", "0 VND", "999,999,999,999 VND", "Số âm / >max"),
            ("Email", "1 char + @ + domain", "RFC 5321", "Không @ / sai domain"),
            ("Ngày", "01/01/1900", "31/12/9999", "32/13/yyyy / chữ"),
        ]),
        ("Tiếng Việt có dấu — test strings", [
            ("Loại", "Sample"),
            ("NFC normalized", "Nguyễn Văn Đỗ"),
            ("NFD decomposed", "Nguyễn Văn Đỗ (decomp)"),
            ("Mixed", "ABC nguyễn xyz Lê 123"),
            ("Special chars", "Q.1, TP.HCM — đường Lê Lợi/Bến Nghé"),
        ]),
    ]
    row = 4
    for title, table in sections:
        ws.merge_cells(f"B{row}:E{row}")
        c = ws[f"B{row}"]; c.value = title
        _apply(c, _header_style(fill=LIGHT_BLUE, font_color=NAVY, size=12))
        row += 1
        for i, hdr in enumerate(table):
            for j, val in enumerate(hdr):
                cell = ws.cell(row, 2 + j, val)
                if i == 0:
                    _apply(cell, _header_style())
                else:
                    _apply(cell, _data_style(wrap=True))
            row += 1
        row += 1


def build_lookups(wb: Workbook):
    """Hidden sheet — single source of truth for all dropdown values."""
    ws = wb.create_sheet("_Tra cứu")
    ws.sheet_state = "hidden"

    cols = [
        ("A", "Index"),
        ("B", "Modules"),
        ("C", "TestTypes"),
        ("D", "Roles"),
        ("E", "Priorities"),
        ("F", "Severities"),
        ("G", "RunStatuses"),
        ("H", "DefectStatuses"),
    ]
    for col, name in cols:
        ws[f"{col}1"].value = name
        ws[f"{col}1"].font = Font(name="Times New Roman", size=11, bold=True)

    data_lists = {
        "B": LOOKUPS["modules"],
        "C": LOOKUPS["test_types"],
        "D": [],  # roles — filled at fill-time per project from actor-registry
        "E": LOOKUPS["priorities"],
        "F": LOOKUPS["severities"],
        "G": LOOKUPS["statuses_run"],
        "H": LOOKUPS["defect_statuses"],
    }
    for col, lst in data_lists.items():
        for i, val in enumerate(lst):
            ws[f"{col}{2 + i}"].value = val


# ─────────── Main ───────────
def build(out_path: Path, project: dict | None = None):
    project = project or {
        "display_name": "[Tên hệ thống]",
        "code": "[Mã DA]",
        "client": "[Khách hàng]",
        "dev_unit": "[Đơn vị phát triển]",
        "description": "[Mô tả phạm vi]",
        "tech_stack": "[Công nghệ]",
        "version": "1.0",
        "today": "dd/mm/yyyy",
        "test_env": "Web (Chrome/Edge/Firefox), backend localhost Docker",
        "test_period": "[Thời gian]",
        "primary_tester": "[Tester chính]",
    }

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_cover(wb, project)
    build_instructions(wb)
    build_overview(wb, project)
    build_dashboard(wb)
    build_tc_sheet_v2(wb, "Tên chức năng", start_data_row=13, end_data=1000)
    build_tc_sheet_v2(wb, "Tên API", start_data_row=13, end_data=1000)
    build_defect_log(wb)
    build_test_data(wb)
    build_lookups(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"✓ Built: {out_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Size: {out_path.stat().st_size / 1024:.1f} KB")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--out", default="C:/Users/James/.claude/skills/generate-docs/engine/templates/test-case.xlsx")
    args = p.parse_args()
    build(Path(args.out))


if __name__ == "__main__":
    main()
